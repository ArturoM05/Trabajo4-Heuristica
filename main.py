"""
Script principal – Trabajo 4: Algoritmos Evolutivos para NWJSSP
================================================================
Ejecuta dos modos según la variable MODE al inicio del archivo:

    MODE = "final"      → Corre la configuración definitiva de cada algoritmo
                          y guarda los Excel de entrega.

    MODE = "parametric" → Corre análisis comparativo de parámetros para DE
                          y para DE-VND, guardando un Excel por configuración.

Algoritmos:
    1. DE      – Evolución Diferencial pura (DE/rand/1 + cruce binomial)
    2. DE-VND  – DE híbrido con VND como refinamiento local
"""

import os
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from constructive import ConstructiveAlgorithm
from de import DESearch
from de_vnd import DEVNDSearch
from read_instances import read_nwjssp_instance

# ===========================================================================
# ════════════════════════  CONFIGURACIÓN PRINCIPAL  ════════════════════════
# ===========================================================================

# Modo de ejecución: "final" | "parametric"
MODE = "final"

# En modo final puede ejecutarse solo un algoritmo si se especifica:
#   FINAL_ALGO_MODE=de       -> solo DE puro
#   FINAL_ALGO_MODE=devnd    -> solo DE-VND
#   FINAL_ALGO_MODE=both     -> ambos (comportamiento por defecto)
FINAL_ALGO_MODE = os.getenv("FINAL_ALGO_MODE", "both").lower()

# Directorio con los archivos .txt de instancias
INSTANCES_DIR       = "instances"
INSTANCES_DIR_PARA  = "instances_parametric"

# Tiempo límite POR EJECUCIÓN de algoritmo
TIME_LIMIT       = 3600   # 1 hora  (modo final)
TIME_LIMIT_PARAM = 300    # 5 min   (modo paramétrico)

# ---------------------------------------------------------------------------
# Parámetros definitivos (usados en MODE="final")
# ---------------------------------------------------------------------------

DE_PARAMS_FINAL = dict(
    NS=20,      # población: equilibrio diversidad/coste
    F=0.8,      # factor de escala clásico
    CR=0.5,     # alta tasa de cruce
    patience=50,
    seed=42,
)

DEVND_PARAMS_FINAL = dict(
    NS=15,              # población ligeramente menor (VND añade coste)
    F=0.8,
    CR=0.9,
    vnd_freq=10,         # VND cada 10 generaciones
    vnd_max_range=10,
    vnd_improve_n1="BI",
    vnd_improve_n2="BI",
    vnd_improve_n3="FI",
    patience=50,
    seed=42,
)

# ---------------------------------------------------------------------------
# Configuraciones paramétricas (usadas en MODE="parametric")
#
# DE puro:   se varía NS (tamaño de población), F (escala), CR (cruce)
#            y patience (convergencia).
# DE-VND:    además se varía vnd_freq (frecuencia del refinamiento VND).
# ---------------------------------------------------------------------------

DE_PARAM_CONFIGS = [
    # ── Variación de NS (tamaño de población) ────────────────────────
    ("DE_NS10_F08_CR09_p50",  dict(NS=10,  F=0.8, CR=0.9, patience=50, seed=42)),
    ("DE_NS20_F08_CR09_p50",  dict(NS=20,  F=0.8, CR=0.9, patience=50, seed=42)),   # <- final
    ("DE_NS30_F08_CR09_p50",  dict(NS=30,  F=0.8, CR=0.9, patience=50, seed=42)),

    # ── Variación de F (factor de escala de mutación) ─────────────────
    ("DE_NS20_F05_CR09_p50",  dict(NS=20,  F=0.5, CR=0.9, patience=50, seed=42)),
    ("DE_NS20_F10_CR09_p50",  dict(NS=20,  F=1.0, CR=0.9, patience=50, seed=42)),

    # ── Variación de CR (tasa de cruce) ───────────────────────────────
    ("DE_NS20_F08_CR05_p50",  dict(NS=20,  F=0.8, CR=0.5, patience=50, seed=42)),
    ("DE_NS20_F08_CR07_p50",  dict(NS=20,  F=0.8, CR=0.7, patience=50, seed=42)),

    # ── Variación de patience (criterio de convergencia) ──────────────
    # Permite ver el efecto del criterio de parada por no mejora
    ("DE_NS20_F08_CR09_p20",  dict(NS=20,  F=0.8, CR=0.9, patience=20,  seed=42)),
    ("DE_NS20_F08_CR09_p100", dict(NS=20,  F=0.8, CR=0.9, patience=100, seed=42)),
]

DEVND_PARAM_CONFIGS = [
    # ── Variación de NS ───────────────────────────────────────────────
    ("DEVND_NS10_F08_CR09_vf5_p50",  dict(NS=10,  F=0.8, CR=0.9, vnd_freq=5,  vnd_max_range=10, vnd_improve_n1="BI", vnd_improve_n2="BI", vnd_improve_n3="FI", patience=50, seed=42)),
    ("DEVND_NS15_F08_CR09_vf5_p50",  dict(NS=15,  F=0.8, CR=0.9, vnd_freq=5,  vnd_max_range=10, vnd_improve_n1="BI", vnd_improve_n2="BI", vnd_improve_n3="FI", patience=50, seed=42)),  # <- final
    ("DEVND_NS20_F08_CR09_vf5_p50",  dict(NS=20,  F=0.8, CR=0.9, vnd_freq=5,  vnd_max_range=10, vnd_improve_n1="BI", vnd_improve_n2="BI", vnd_improve_n3="FI", patience=50, seed=42)),

    # ── Variación de vnd_freq (frecuencia del refinamiento VND) ──────
    ("DEVND_NS15_F08_CR09_vf2_p50",  dict(NS=15,  F=0.8, CR=0.9, vnd_freq=2,  vnd_max_range=10, vnd_improve_n1="BI", vnd_improve_n2="BI", vnd_improve_n3="FI", patience=50, seed=42)),
    ("DEVND_NS15_F08_CR09_vf10_p50", dict(NS=15,  F=0.8, CR=0.9, vnd_freq=10, vnd_max_range=10, vnd_improve_n1="BI", vnd_improve_n2="BI", vnd_improve_n3="FI", patience=50, seed=42)),

    # ── Variación de F ────────────────────────────────────────────────
    ("DEVND_NS15_F05_CR09_vf5_p50",  dict(NS=15,  F=0.5, CR=0.9, vnd_freq=5,  vnd_max_range=10, vnd_improve_n1="BI", vnd_improve_n2="BI", vnd_improve_n3="FI", patience=50, seed=42)),
    ("DEVND_NS15_F10_CR09_vf5_p50",  dict(NS=15,  F=1.0, CR=0.9, vnd_freq=5,  vnd_max_range=10, vnd_improve_n1="BI", vnd_improve_n2="BI", vnd_improve_n3="FI", patience=50, seed=42)),

    # ── Variación de patience ─────────────────────────────────────────
    ("DEVND_NS15_F08_CR09_vf5_p20",  dict(NS=15,  F=0.8, CR=0.9, vnd_freq=5,  vnd_max_range=10, vnd_improve_n1="BI", vnd_improve_n2="BI", vnd_improve_n3="FI", patience=20,  seed=42)),
    ("DEVND_NS15_F08_CR09_vf5_p100", dict(NS=15,  F=0.8, CR=0.9, vnd_freq=5,  vnd_max_range=10, vnd_improve_n1="BI", vnd_improve_n2="BI", vnd_improve_n3="FI", patience=100, seed=42)),
]

# ===========================================================================


# ---------------------------------------------------------------------------
# Utilidades Excel (idénticas al trabajo anterior)
# ---------------------------------------------------------------------------

def create_results_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


def get_column_letter(col_num):
    col_letter = ""
    col = col_num
    while col >= 0:
        col_letter = chr(65 + (col % 26)) + col_letter
        col = col // 26 - 1
        if col < 0:
            break
    return col_letter


def add_results_sheet(workbook, instance_name, flow_time, computation_time, job_start_times):
    ws = workbook.create_sheet(instance_name)
    ws["A1"] = int(flow_time)
    ws["B1"] = int(round(computation_time))

    for idx, start_time in enumerate(job_start_times):
        ws[f"{get_column_letter(idx)}2"] = int(start_time)

    fill  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font  = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    for cell in [ws["A1"], ws["B1"]]:
        cell.fill  = fill
        cell.font  = font
        cell.alignment = align

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    for idx in range(len(job_start_times)):
        ws.column_dimensions[get_column_letter(idx)].width = 12


# ---------------------------------------------------------------------------
# Helpers de ejecución
# ---------------------------------------------------------------------------

def _initial_solution(n, m, operations, release_dates):
    c = ConstructiveAlgorithm(n, m, operations, release_dates)
    sol, _, _ = c.solve()
    return sol


def run_de(n, m, operations, release_dates, time_limit=TIME_LIMIT, **params):
    sol0  = _initial_solution(n, m, operations, release_dates)
    algo  = DESearch(n, m, operations, release_dates, time_limit=time_limit, **params)
    return algo.solve(initial_solution=sol0)   # (starts, flow, comp, n_evals)


def run_devnd(n, m, operations, release_dates, time_limit=TIME_LIMIT, **params):
    sol0  = _initial_solution(n, m, operations, release_dates)
    algo  = DEVNDSearch(n, m, operations, release_dates, time_limit=time_limit, **params)
    return algo.solve(initial_solution=sol0)


def process_instance(instance_file, algo_fn, time_limit, params):
    try:
        n, m, ops, rd, _ = read_nwjssp_instance(instance_file)
        starts, ft, ct, n_evals = algo_fn(n, m, ops, rd, time_limit=time_limit, **params)
        name = os.path.splitext(os.path.basename(instance_file))[0]
        return name, starts, ft, ct, n_evals
    except Exception as e:
        print(f"  Error procesando {instance_file}: {e}")
        import traceback; traceback.print_exc()
        return None


def run_batch(instance_files, algo_fn, label, output_file, time_limit, params):
    """Ejecuta algo_fn sobre todas las instancias y guarda Excel."""
    print(f"\n  [{label}]")
    print("  " + "-" * 72)

    wb          = create_results_workbook()
    total_time  = 0
    total_evals = 0
    count       = 0

    for f in instance_files:
        result = process_instance(f, algo_fn, time_limit, params)
        if result:
            name, sol, ft, ct, n_evals = result
            ev_str = f"{n_evals:8d}" if n_evals is not None else "     N/A"
            print(f"    {name:28s} | Z={ft:12.0f} | t={ct:10.2f}ms | evals={ev_str}")
            total_time  += ct
            total_evals += n_evals if n_evals is not None else 0
            count       += 1
            add_results_sheet(wb, name, ft, ct, sol)

    wb.save(output_file)
    avg      = total_time  / count if count > 0 else 0
    avg_eval = total_evals / count if count > 0 else 0
    print(f"\n  ✓ {output_file}  "
          f"(inst={count}, total={total_time/1000:.2f}s, "
          f"avg={avg:.2f}ms, avg_evals={avg_eval:.1f})")
    return total_time, count


# ---------------------------------------------------------------------------
# Modo FINAL
# ---------------------------------------------------------------------------

def run_final(instance_files):
    print("\n" + "=" * 75)
    print("MODO FINAL  –  Configuraciones definitivas")
    print(f"Algoritmo(s): {FINAL_ALGO_MODE.upper()}")
    print("=" * 75)

    all_configs = {
        "de":    (run_de,    "DE  (NS=20, F=0.8, CR=0.9)",
                  "NWJSSP_ArturoMurgueytio_DE.xlsx",    TIME_LIMIT, DE_PARAMS_FINAL),
        "devnd": (run_devnd, "DE-VND  (NS=15, F=0.8, CR=0.9, vnd_freq=5)",
                  "NWJSSP_ArturoMurgueytio_DEVND.xlsx", TIME_LIMIT, DEVND_PARAMS_FINAL),
    }

    if FINAL_ALGO_MODE in ("both", "all", ""):
        configs = list(all_configs.values())
    elif FINAL_ALGO_MODE in all_configs:
        configs = [all_configs[FINAL_ALGO_MODE]]
    else:
        print(f"Advertencia: FINAL_ALGO_MODE='{FINAL_ALGO_MODE}' no reconocido. Ejecutando ambos.")
        configs = list(all_configs.values())

    summary = []
    for fn, label, outfile, tl, params in configs:
        total, count = run_batch(instance_files, fn, label, outfile, tl, params)
        summary.append((label, count, total))

    _print_summary(summary)
    return summary


# ---------------------------------------------------------------------------
# Modo PARAMÉTRICO
# ---------------------------------------------------------------------------

def run_parametric(instance_files):
    print("\n" + "=" * 75)
    print("MODO PARAMÉTRICO  –  Análisis comparativo de parámetros")
    print(f"Tiempo límite por configuración: {TIME_LIMIT_PARAM} s")
    print("=" * 75)

    all_results = []

    # ── DE puro ──────────────────────────────────────────────────────
    print("\n▸ DE  –  variando NS, F y CR")
    for label, params in DE_PARAM_CONFIGS:
        outfile = f"NWJSSP_ArturoMurgueytio_{label}.xlsx"
        total, count = run_batch(
            instance_files, run_de, label, outfile, TIME_LIMIT_PARAM, params
        )
        all_results.append((label, count, total))

    # ── DE-VND ───────────────────────────────────────────────────────
    print("\n▸ DE-VND  –  variando NS, vnd_freq y F")
    for label, params in DEVND_PARAM_CONFIGS:
        outfile = f"NWJSSP_ArturoMurgueytio_{label}.xlsx"
        total, count = run_batch(
            instance_files, run_devnd, label, outfile, TIME_LIMIT_PARAM, params
        )
        all_results.append((label, count, total))

    _print_summary(all_results)
    return all_results


# ---------------------------------------------------------------------------
# Resumen impreso
# ---------------------------------------------------------------------------

def _print_summary(results):
    print("\n" + "=" * 75)
    print("RESUMEN")
    print("=" * 75)
    print(f"  {'Configuración':<45} | {'Inst.':>5} | {'Total':>9} | {'Promedio':>10}")
    print("  " + "-" * 74)
    for label, count, total in results:
        avg = total / count if count > 0 else 0
        print(f"  {label:<45} | {count:>5} | {total/1000:>6.2f}s  | {avg:>8.2f}ms")
    print()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 75)
    print("NWJSSP – Trabajo 4: DE  |  DE-VND (Algoritmos Evolutivos)")
    print(f"Modo: {MODE.upper()}")
    print("=" * 75)

    if MODE == "final":
        instance_files = sorted(glob.glob(os.path.join(INSTANCES_DIR, "*.txt")))
        if not instance_files:
            print(f"Error: No se encontraron instancias en '{INSTANCES_DIR}'")
            return
        print(f"\nInstancias encontradas: {len(instance_files)}")
        run_final(instance_files)

    elif MODE == "parametric":
        instance_files = sorted(glob.glob(os.path.join(INSTANCES_DIR_PARA, "*.txt")))
        if not instance_files:
            print(f"Error: No se encontraron instancias en '{INSTANCES_DIR_PARA}'")
            return
        print(f"\nInstancias encontradas: {len(instance_files)}")
        run_parametric(instance_files)

    else:
        print(f"Error: MODE='{MODE}' no reconocido. Use 'final' o 'parametric'.")
        return

    print("✓ Ejecución completada\n")


if __name__ == "__main__":
    main()